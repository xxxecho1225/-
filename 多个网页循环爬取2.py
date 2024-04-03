from __future__ import annotations
import sys, os, time, base64, signal, logging, traceback, rpdb
from argparse import ArgumentParser
from datetime import datetime as dt, timedelta as td
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
from openpyxl import Workbook
import copy
from openpyxl import load_workbook

bind_address = "0.0.0.0"
wait_timeout = 10
poll_frequency = 0.2
tmp_dir = "/tmp"

logging_level = "DEBUG"
logging_format = "%(asctime)s --- %(levelname)s: %(message)s"


os.environ["CUDA_VISIBLE_DEVICES"] = "-1"
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"
for proxy in ["https_proxy", "http_proxy", "socks_proxy", "all_proxy"]:
    os.environ[proxy] = os.environ[proxy.upper()] = ""

_driver: WebDriver
_logger: logging.Logger

urls = []
university_names = []  # 存储大学名称的列表

def read():
    # 读取 Excel 文件
    workbook = load_workbook(filename='D:/git_python/school/数据分析/学校分数信息/src/b.xlsx')

    # 获取第一个工作表
    sheet = workbook.active

    # 循环处理每一行
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        university_name = row[0]  # 第一列是大学名称
        url = row[1]  # 第二列是大学地址
        urls.append(url)
        #print(university_name)
        #print(url)
        university_names.append(university_name)
    return university_names
read()

def restart_session(display) -> None:
    global _driver
    _driver = new_session(display)

def wait_find(driver: WebDriver, by: str, value: str, all=True):
    if all:
        finder = lambda d: d.find_elements(by, value)
    else:
        finder = lambda d: d.find_element(by, value)
    return WebDriverWait(driver, wait_timeout, poll_frequency).until(finder)

def new_session(display) -> WebDriver:
    options = FirefoxOptions()
    if display == 'headless':
        options.headless = True
    # 设置 Firefox 可执行文件的路径
    firefox_binary = "C:/Program Files/Mozilla Firefox/firefox.exe"
    options.binary_location = firefox_binary
    driver = webdriver.Firefox(options=options)
    return driver

def get_provinces(driver):
    for url, university_name in zip(urls,university_names):
        driver.get(url)
        crawl_data(driver, university_name)
        
        # 在处理完一个 URL 后重新启动会话
        driver = new_session('headless')

def crawl_data(driver,university_name):
    # 定位省份元素并循环
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="app"]/div/div/div/div/div/div/div/div/div/div')))
    province_elements = driver.find_elements(By.XPATH, '//*[@id="app"]/div/div/div/div/div/div/div/div[1]/div[2]/div/div')
    for province_element in province_elements:
        province_name = province_element.text
        province_element.click()
        print("省份名字:",province_name)
        
        # 定位年份元素并循环
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="app"]/div/div/div/div/div/div/div/div/div/div/div')))
        year_elements = driver.find_elements(By.XPATH, '//*[@id="app"]/div/div/div/div/div/div/div/div[2]/div/div/div')
        for year_element in year_elements:
            year_name = year_element.text
            year_element.click()
            print("年份名字：",year_name)

            data = get_table_data(driver)
            saveIntoCsv(data,province_name, year_name,university_name, driver)
    
    driver.quit()

def saveIntoCsv(data, province_name, year_name, university_name,driver):
    wb = Workbook()
    # 首先删除默认创建的工作表
    ws = wb.active
    
    # 写入表头数据
    header_data = get_table_header(driver)
    ws.append(header_data)
    
    # 写入表格数据
    for row in data:
        ws.append(row)

    # 保存工作簿到文件
    filename = f"{year_name}-{province_name}-{university_name}.csv"
    os.makedirs(f'学校信息/{university_name}', exist_ok=True)
    wb.save(os.path.join(f'学校信息/{university_name}', filename))

def clean_sheet_title(title):
    # 使用 replace 方法过滤掉斜杠
    return title.replace('/', '')

def get_table_header(driver):
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'el-table__header')))
    
    # 找到表头元素
    header = driver.find_element(By.CLASS_NAME, 'el-table__header')
    # 找到所有的列头元素
    column_headers = header.find_elements(By.XPATH, '//*[@id="app"]/div/div/div/div/div/div/div/div/table/thead/tr/th/div')

    
    # 提取列头文本
    header_data = [column_header.text.strip() for column_header in column_headers]
    
    print("表格头部信息：", header_data)
    
    return header_data

def get_table_data(driver):
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'el-table__body-wrapper')))
    
    # 找到表格数据包装器元素
    data_wrapper = driver.find_element(By.CLASS_NAME, 'el-table__body-wrapper')
    
    # 等待表格数据加载完成
    WebDriverWait(data_wrapper, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'el-table__body')))
    
    # 找到表格数据元素
    table_body = data_wrapper.find_element(By.CLASS_NAME, 'el-table__body')

    #WebDriverWait(data_wrapper, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="app"]/div/div/div/div/div/div/div/div/table/tbody/tr')))
    
    # 找到所有行数据
    rows = table_body.find_elements(By.XPATH, '//*[@id="app"]/div/div/div/div/div/div/div/div/table/tbody/tr')
    data = []

 # 提取每行数据
    for row in rows:
        try:
            # 找到当前行下的所有td元素
            cells = row.find_elements(By.XPATH, './td')
            row_data = []
            # 提取每个td中的div的文本内容
            for cell in cells:
                div = cell.find_element(By.XPATH, './div')
                row_data.append(div.text.strip())
            data.append(row_data)
        except StaleElementReferenceException:
            # 如果发生异常，则重新定位元素
            rows = table_body.find_elements(By.XPATH, '//*[@id="app"]/div/div/div/div/div/div/div/div/table/tbody/tr')
            continue

    print("表格数据：", data)

    
    return data



def run(logger: logging.Logger):
    global _logger
    _logger = logger
    try:
        restart_session('headless')
        provinces = get_provinces(_driver)
        logger.info("success")
    except KeyboardInterrupt:
        os.abort()
    except RuntimeError:
        rpdb.set_trace(bind_address, 0)
        return
    except:
        logger.critical(traceback.format_exc())
        # logger.info(str(locals()))
        rpdb.set_trace(bind_address, 0)

if __name__ == "__main__":
    logger = logging.Logger("main", logging_level)
    handler = logging.StreamHandler()
    handler.setFormatter(logging.Formatter(logging_format))
    logger.addHandler(handler)
    run(logger)
