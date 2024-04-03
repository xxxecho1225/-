#清华镜像 https://pypi.tuna.tsinghua.edu.cn/simple

from openpyxl import load_workbook

# 读取 Excel 文件
workbook = load_workbook(filename='D:/git_python/school/数据分析/学校分数信息/src/a.xlsx')

# 获取第一个工作表
sheet = workbook.active

# 循环处理每一行
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
    university_name = row[0]  # 第一列是大学名称
    university_address = row[1]  # 第二列是大学地址
    print(university_name)
    print(university_address)


