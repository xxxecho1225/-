from __future__ import annotations
import sys, os, time, base64, signal, logging, traceback, rpdb
from argparse import ArgumentParser
from datetime import datetime as dt, timedelta as td
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
from openpyxl import Workbook
import copy
from openpyxl import load_workbook

bind_address = "0.0.0.0"
wait_timeout = 10
poll_frequency = 0.2
tmp_dir = "/tmp"

logging_level = "DEBUG"
logging_format = "%(asctime)s --- %(levelname)s: %(message)s"


os.environ["CUDA_VISIBLE_DEVICES"] = "-1"
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"
for proxy in ["https_proxy", "http_proxy", "socks_proxy", "all_proxy"]:
    os.environ[proxy] = os.environ[proxy.upper()] = ""

_driver: WebDriver
_logger: logging.Logger

urls = []
university_names = []  # 存储大学名称的列表

def read():
    # 读取 Excel 文件
    workbook = load_workbook(filename='D:/git_python/school/数据分析/学校分数信息/src/a.xlsx')

    # 获取第一个工作表
    sheet = workbook.active

    # 循环处理每一行
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        university_name = row[0]  # 第一列是大学名称
        url = row[1]  # 第二列是大学地址
        urls.append(url)
        #print(university_name)
        #print(url)
        university_names.append(university_name)
    return university_names
read()

def restart_session(display) -> None:
    global _driver
    _driver = new_session(display)

def wait_find(driver: WebDriver, by: str, value: str, all=True):
    if all:
        finder = lambda d: d.find_elements(by, value)
    else:
        finder = lambda d: d.find_element(by, value)
    return WebDriverWait(driver, wait_timeout, poll_frequency).until(finder)

def new_session(display) -> WebDriver:
    options = FirefoxOptions()
    if display == 'headless':
        options.headless = True
    # 设置 Firefox 可执行文件的路径
    firefox_binary = "C:/Program Files/Mozilla Firefox/firefox.exe"
    options.binary_location = firefox_binary
    driver = webdriver.Firefox(options=options)
    return driver

def get_provinces(driver):
    for url, university_name in zip(urls,university_names):
        driver.get(url)
        crawl_data(driver, university_name)

def crawl_data(driver, university_name):
    data_all = {}
    WebDriverWait(driver, 200).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'dd[data-param="ssmc"]')))
    province_elements = driver.find_elements(By.CSS_SELECTOR, 'dd[data-param="ssmc"] > a')
    province_len = len(province_elements)
    for province_id in range(province_len):
        province_elements = driver.find_elements(By.CSS_SELECTOR, 'dd[data-param="ssmc"] > a')
        province_name = province_elements[province_id].get_attribute('data-value')
        print('省份：', province_name)
        province_elements[province_id].click()
        
        WebDriverWait(driver, 200).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'dd[data-param="zsnf"]')))
        year_elements = driver.find_elements(By.CSS_SELECTOR, 'dd[data-param="zsnf"] > a')
        year_len = len(year_elements)
        for year_id in range(year_len):
            year_elements = driver.find_elements(By.CSS_SELECTOR, 'dd[data-param="zsnf"] > a')
            year_name = year_elements[year_id].get_attribute('data-value')
            print('年份：', year_name)
            year_elements[year_id].click()
            
            WebDriverWait(driver, 200).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'dd[data-param="klmc"]')))
            subject_elements = driver.find_elements(By.CSS_SELECTOR, 'dd[data-param="klmc"] > a')
            subject_len = len(subject_elements)
            
            
            for subject_id in range(subject_len):
                subject_elements = driver.find_elements(By.CSS_SELECTOR, 'dd[data-param="klmc"] > a')
                if subject_id >= len(subject_elements):
                        continue  # 跳过超出列表长度的索引
                subject_name = subject_elements[subject_id].get_attribute('data-value')
                print('科类：', subject_name)
                subject_elements[subject_id].click()
                
                WebDriverWait(driver, 200).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'dd[data-param="zslx"]')))
                type_elements = driver.find_elements(By.CSS_SELECTOR, 'dd[data-param="zslx"] > a')
                type_len = len(type_elements)
                for type_id in range(type_len):
                    type_elements = driver.find_elements(By.CSS_SELECTOR, 'dd[data-param="zslx"] > a')
                    type_name = type_elements[type_id].get_attribute('data-value')
                    print('类型：', type_name)
                    type_elements[type_id].click()
                    
                    data = get_table_data(driver)
                    #工作表格式
                    data_all[f'{subject_name}_{type_name}'] = copy.deepcopy(data)
            saveIntoCsv(data_all, province_name, year_name,university_name)

def saveIntoCsv(data, province_name, year_name, university_name):
    wb = Workbook()
    # 首先删除默认创建的工作表
    default_ws = wb.active
    wb.remove(default_ws)
    for subject_type_key, rows in data.items():
        ws = wb.create_sheet(title=clean_sheet_title(subject_type_key))
        for row in rows:
            ws.append(row)
    # 保存工作簿到文件
    filename = f"{year_name}-{province_name}-{university_name}.csv"
    os.makedirs(f'学校信息/{university_name}', exist_ok=True)
    wb.save(os.path.join(f'学校信息/{university_name}', filename))

def clean_sheet_title(title):
    # 使用 replace 方法过滤掉斜杠
    return title.replace('/', '')

def get_table_data(driver):
    WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.ID, 'sszygradeListPlace')))
    table = driver.find_element(By.ID, 'sszygradeListPlace')
    # 使用XPath选择所有td和th元素
    data = []
    i = 0
    while True:
        try:
            row = table.find_elements(By.XPATH, './/tr')[i]
            cells = row.find_elements(By.XPATH, './/td | .//th')
            row_data = [cell.get_attribute('innerText') for cell in cells]
            if row_data:
                data.append(row_data)
            i += 1
        except IndexError:
            break
        except StaleElementReferenceException:
            # 页面元素已过时，尝试重新查找表格元素
            table = driver.find_element(By.ID, 'sszygradeListPlace')
            continue

    return data



def run(logger: logging.Logger):
    global _logger
    _logger = logger
    try:
        restart_session('headless')
        provinces = get_provinces(_driver)
        logger.info("success")
    except KeyboardInterrupt:
        os.abort()
    except RuntimeError:
        rpdb.set_trace(bind_address, 0)
        return
    except:
        logger.critical(traceback.format_exc())
        # logger.info(str(locals()))
        rpdb.set_trace(bind_address, 0)

if __name__ == "__main__":
    logger = logging.Logger("main", logging_level)
    handler = logging.StreamHandler()
    handler.setFormatter(logging.Formatter(logging_format))
    logger.addHandler(handler)
    run(logger)
